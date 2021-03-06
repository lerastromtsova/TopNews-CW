import csv
import openpyxl
from math import ceil
from coefs import COEFFICIENT_1_FOR_NEWS, COEFFICIENT_2_FOR_NEWS


def check_xl(nodes, all_links):
    file = csv.writer(open("t.csv", "w"), delimiter=',')

    for n in nodes:
        countries = {d.country for d in n.documents}
        row = [f"{n.name} | {countries}"]
        for k, m in all_links[n].items():
            row.append(f"{k.name} | {m} | {' '.join({d.country for d in k.documents})}")
            cd = set(n.documents).intersection(set(k.documents))
            cc = {d.country for d in cd}
            perc = len(cc)/len(countries)
            row.append(perc)
            row.append(cc)
            tex = ' | '.join([d.title for d in cd])
            row.append(tex)
        file.writerow(row)


def write_topics_with_freq(fname, nodes, with_children=False):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = 'Topic'
    sheet['B1'] = 'News'
    sheet['C1'] = 'Keywords'
    for i, n in enumerate(nodes):
        if with_children:
            text = ''

            if n.subtopics:
                for c in n.subtopics:

                    text += ' '.join(c.name)
                    text += ' | '
            else:
                if type(n.name) == 'str':
                    text = n.name
                else:
                    text = ' '.join(n.name)
        else:
            if type(n.name) == 'str':
                text = n.name
            else:
                text = ' '.join(n.name)
        freq_words = n.find_most_frequent(type="uppercase")
        for word, val in freq_words.items():
            t = f"|{word}:{val}|"
            text += t
        sheet.cell(row=i+2, column=1).value = text
        docs = n.documents
        for j,doc in enumerate(docs):
            sheet.cell(row=i+2, column=j+2).value = f"{doc.country} | {doc.translated['title']} | {doc.translated['lead']} | {doc.url} | {doc.translated['content']} | {doc.named_entities['title']} | {doc.named_entities['content']}"
    wb.save(fname)


def get_subtopics_string(node):
    if node.subtopics:
        return '|'.join([' '.join(s.name) for s in node.subtopics])
    return None


def write_topics_to_xl(fname, nodes, with_children=False):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = 'Topic'
    sheet['B1'] = 'News'
    sheet['C1'] = 'Keywords'
    for i, n in enumerate(nodes):
        if with_children:
            text = ''
            if n.subtopics:
                print(len(n.subtopics))
                for c in n.subtopics:
                    text += ' '.join(c.name)
                    text += ' | '
            else:
                if type(n.name) == 'str':
                    text = n.name
                else:
                    text = ' '.join(n.name)
        else:
            if type(n.name) == 'str':
                text = n.name
            else:
                text = ' '.join(n.name)
        sheet.cell(row=i+2, column=1).value = text
        docs = n.documents
        for j,doc in enumerate(docs):
            sheet.cell(row=i+2, column=j+2).value = f"{doc.country} | {doc.translated['title']} | {doc.translated['lead']} | {doc.url} | {doc.translated['content']} | {doc.named_entities['title']} | {doc.named_entities['content']}"
    wb.save(fname)


def write_tz(tree):
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i,node in enumerate(tree.roots):
        texts = get_all_in_strings(node, [])
        for j, text in enumerate(texts):
            print(text)
            sheet.cell(row=i+1, column=j+1).value = text
    wb.save("tz-today.xlsx")


def get_all_in_strings(node, prev):
    prev.append(f"{node.name} | {len(node.documents)}")
    for ch in node.children:
        if ch.children:
            get_all_in_strings(ch, prev)
    return prev


def write_words_to_xl(fname, data):
    file = csv.writer(open(fname, "w"), delimiter=',')
    headers = ['Document', 'Deleted words']
    file.writerow(headers)

    for r in data:
        row = [r.translated, r.removed_words]
        file.writerow(row)


def write_start_words_to_xl(fname, data):
    file = csv.writer(open(fname, "w"))
    headers = ['Document', 'Deleted words']
    file.writerow(headers)

    for r in data:
        for key, ind in r.start_words.items():
            row = [key, ind]
            file.writerow(row)

def write_rows_content(fname, similarities):
    wb = openpyxl.Workbook()
    sheet = wb.active
    i = 1
    for row in similarities:
        j = 1
        # setlist = [doc.description for doc in row]
        setlist = [doc.named_entities['content'] for doc in row]
        topic = set.intersection(*setlist)
        sheet.cell(row=i,column=j).value = ' '.join(topic)
        for doc in row:
            j += 1
            sheet.cell(row=i,column=j).value = f"|{doc.id}| {doc.country} | {doc.url} | {doc.translated['title']} | " \
                                               f"{doc.named_entities['title']} | {doc.translated['lead']} |" \
                                               f"{doc.translated['content']} | {doc.named_entities['content']}"
        i += 1
    wb.save(fname)

def write_rows_title(fname, similarities):
    wb = openpyxl.Workbook()
    sheet = wb.active
    i = 1
    for row in similarities:
        j = 1
        # setlist = [doc.description for doc in row]
        setlist = [{word.lower() for word in doc.tokens['title']} for doc in row]
        topic = set.intersection(*setlist)
        sheet.cell(row=i,column=j).value = ' '.join(topic)
        for doc in row:
            j += 1
            sheet.cell(row=i,column=j).value = f"|{doc.id}| {doc.country} | {doc.url} | {doc.translated['title']} | " \
                                               f"{doc.named_entities['title']} | {doc.translated['lead']} |" \
                                               f"{doc.translated['content']} | {doc.named_entities['content']}"
        i += 1
    wb.save(fname)

def write_topics(fname, topics):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.cell(row=1, column=1).value = "Method"
    sheet.cell(row=1, column=2).value = "Old name"  # should be empty
    sheet.cell(row=1, column=3).value = "Name"  # should be empty
    sheet.cell(row=1, column=4).value = "Topic"
    sheet.cell(row=1, column=5).value = "All unique words"

    sheet.cell(row=1, column=6).value = "News1"
    sheet.cell(row=1, column=7).value = "News2"

    sheet.cell(row=1, column=8).value = "# All"
    sheet.cell(row=1, column=9).value = "# of unique"

    sheet.cell(row=1, column=10).value = "# of All-countries"
    sheet.cell(row=1, column=11).value = "# of All-countries-small"
    sheet.cell(row=1, column=12).value = "# of FIO"
    sheet.cell(row=1, column=13).value = "# of Big letter"
    sheet.cell(row=1, column=14).value = "# of Small letter"
    sheet.cell(row=1, column=15).value = "# of Countries"
    sheet.cell(row=1, column=16).value = "# of Numbers"
    sheet.cell(row=1, column=17).value = "# of IDs"
    sheet.cell(row=1, column=18).value = "Frequent Unique"
    sheet.cell(row=1, column=19).value = "u# of All-countries (unique)"
    sheet.cell(row=1, column=20).value = "u# of unique"
    sheet.cell(row=1, column=21).value = "u# of All-countries-small (unique)"
    sheet.cell(row=1, column=22).value = "u# of FIO (unique)"
    sheet.cell(row=1, column=23).value = "u# of Big letter (unique)"
    sheet.cell(row=1, column=24).value = "u# of Small letter (unique)"
    sheet.cell(row=1, column=25).value = "u# of Countries (unique)"
    sheet.cell(row=1, column=26).value = "uNumbers (unique)"
    sheet.cell(row=1, column=27).value = "uIDs (unique)"
    sheet.cell(row=1, column=28).value = "Small letter"
    sheet.cell(row=1, column=29).value = "Countries"
    sheet.cell(row=1, column=30).value = "Numbers"
    sheet.cell(row=1, column=31).value = "IDs"
    sheet.cell(row=1, column=32).value = "uAll unique words"
    sheet.cell(row=1, column=33).value = "uAll-countries (unique)"
    sheet.cell(row=1, column=34).value = "uAll-countries-small (unique)"
    sheet.cell(row=1, column=35).value = "uFIO (unique)"
    sheet.cell(row=1, column=36).value = "uBig letter (unique)"
    sheet.cell(row=1, column=37).value = "uSmall letter (unique)"
    sheet.cell(row=1, column=38).value = "uCountries (unique)"
    sheet.cell(row=1, column=39).value = "u# of Numbers (unique)"
    sheet.cell(row=1, column=40).value = "u# of IDs (unique)"
    sheet.cell(row=1, column=41).value = "Most frequent"
    sheet.cell(row=1, column=42).value = "50% frequent"
    sheet.cell(row=1, column=43).value = "What"
    sheet.cell(row=1, column=44).value = "What2"
    sheet.cell(row=1, column=45).value = "Common in descriptions"
    sheet.cell(row=1, column=46).value = "Common in text"
    sheet.cell(row=1, column=47).value = "Comment"  # should be empty
    sheet.cell(row=1, column=48).value = "Name"
    sheet.cell(row=1, column=49).value = "Topic"

    sheet.cell(row=1, column=50).value = "FIO coef 1"
    sheet.cell(row=1, column=51).value = "Big coef 1"
    sheet.cell(row=1, column=52).value = "Small coef 1"
    sheet.cell(row=1, column=53).value = "Countries coef 1"

    sheet.cell(row=1, column=54).value = "FIO coef Y"
    sheet.cell(row=1, column=55).value = "Big coef Y"
    sheet.cell(row=1, column=56).value = "Small coef Y"
    sheet.cell(row=1, column=57).value = "Countries coef Y"

    sheet.cell(row=1, column=58).value = "FIO coef 2"
    sheet.cell(row=1, column=59).value = "Big coef 2"
    sheet.cell(row=1, column=60).value = "Small coef 2"
    sheet.cell(row=1, column=61).value = "Countries coef 2"

    sheet.cell(row=1, column=62).value = "ID coef"

    sheet.cell(row=1, column=63).value = "Sum 1"
    sheet.cell(row=1, column=64).value = "Sum 2"
    sheet.cell(row=1, column=65).value = "F"

    sheet.cell(row=1, column=66).value = "# of added news, their ids"
    sheet.cell(row=1, column=67).value = "News"
    sheet.cell(row=1, column=68).value = "Method"

    # counts = [0]*7

    for i, topic in enumerate(topics):

        all_words, _ = topic.all_words(topic.name)
        all_wo_countries, num_all_wo_countries = topic.all_wo_countries(all_words)
        all_wo_countries_small, num_all_wo_countries_small = topic.all_wo_countries_and_small(all_words)
        fio, num_fio = topic.fio(all_words)
        big, num_big = topic.big(all_words)
        small, num_small = topic.small(all_words)
        countries, num_countries = topic.countries(all_words)
        numbers, num_numbers = topic.numbers(all_words)
        ids, num_ids = topic.ids(all_words)

        unique_words, num_unique_words = topic.all_words(topic.new_name)
        unique_wo_countries, num_unique_wo_countries = topic.all_wo_countries(unique_words)
        unique_wo_countries_small, num_unique_wo_countries_small = topic.all_wo_countries_and_small(unique_words)
        unique_fio, unique_num_fio = topic.fio(unique_words)
        unique_big, unique_num_big = topic.big(unique_words)
        unique_small, unique_num_small = topic.small(unique_words)
        unique_countries, unique_num_countries = topic.countries(unique_words)
        unique_numbers, unique_num_numbers = topic.numbers(unique_words)
        unique_ids, unique_num_ids = topic.ids(unique_words)
        frequent_unique = unique_words.intersection(topic.most_frequent(filename='18'+fname))

        if topic.subtopics:
            name = ''
            for s in topic.subtopics:
                name += ', '.join(s.name)
                # name += f"{', '.join(s.name)} | {s.text_name} |"
        else:
            name = ', '.join(topic.name)
            # name = f"{', '.join(topic.name)} | {topic.text_name}"

        sheet.cell(row=i + 3, column=1).value = " ".join(topic.method)
        sheet.cell(row=i + 3, column=2).value = ", ".join(topic.old_name)
        sheet.cell(row=i + 3, column=3).value = ""
        if topic.subtopics:
            name = ''
            for s in topic.subtopics:
                name += ', '.join(s.name)
                name += "|"
        else:
            name = ", ".join(topic.name)

        sheet.cell(row=i + 3, column=4).value = name
        sheet.cell(row=i + 3, column=5).value = ', '.join(unique_words)
        doc = topic.news[0]
        sheet.cell(row=i + 3, column=6).value = f"|{doc.id}| {doc.country} | {doc.url} | {doc.translated['title']} | " \
                                                f"{doc.translated['lead']} | " \
                                                f"{doc.translated['content']} | Из краткого: {doc.description} | Из текста: {doc.named_entities['content']}"
        doc = topic.news[1]
        sheet.cell(row=i + 3, column=7).value = f"|{doc.id}| {doc.country} | {doc.url} | {doc.translated['title']} | " \
                                                f"{doc.translated['lead']} | " \
                                                f"{doc.translated['content']} | Из краткого: {doc.description} | Из текста: {doc.named_entities['content']}"

        sheet.cell(row=i + 3, column=8).value = len(all_words)
        sheet.cell(row=i + 3, column=9).value = len(unique_words)

        sheet.cell(row=i + 3, column=10).value = num_all_wo_countries
        sheet.cell(row=i + 3, column=11).value = num_all_wo_countries_small
        sheet.cell(row=i + 3, column=12).value = num_fio
        sheet.cell(row=i + 3, column=13).value = num_big
        sheet.cell(row=i + 3, column=14).value = num_small
        sheet.cell(row=i + 3, column=15).value = num_countries
        sheet.cell(row=i + 3, column=16).value = num_numbers
        sheet.cell(row=i + 3, column=17).value = num_ids

        sheet.cell(row=i + 3, column=18).value = ", ".join(frequent_unique)
        sheet.cell(row=i + 3, column=19).value = num_unique_wo_countries
        sheet.cell(row=i + 3, column=20).value = num_unique_words
        sheet.cell(row=i + 3, column=21).value = num_unique_wo_countries_small
        sheet.cell(row=i + 3, column=22).value = unique_num_fio
        sheet.cell(row=i + 3, column=23).value = unique_num_big
        sheet.cell(row=i + 3, column=24).value = unique_num_small
        sheet.cell(row=i + 3, column=25).value = unique_num_countries
        sheet.cell(row=i + 3, column=26).value = ", ".join(unique_numbers)
        sheet.cell(row=i + 3, column=27).value = ", ".join(unique_ids)
        sheet.cell(row=i + 3, column=28).value = ", ".join(small)
        sheet.cell(row=i + 3, column=29).value = ", ".join(countries)
        sheet.cell(row=i + 3, column=30).value = ", ".join(numbers)
        sheet.cell(row=i + 3, column=31).value = ", ".join(ids)
        sheet.cell(row=i + 3, column=32).value = ", ".join(unique_words)
        sheet.cell(row=i + 3, column=33).value = ", ".join(unique_wo_countries)
        sheet.cell(row=i + 3, column=34).value = ", ".join(unique_wo_countries_small)
        sheet.cell(row=i + 3, column=35).value = ", ".join(unique_fio)
        sheet.cell(row=i + 3, column=36).value = ", ".join(unique_big)
        sheet.cell(row=i + 3, column=37).value = ", ".join(unique_small)
        sheet.cell(row=i + 3, column=38).value = ", ".join(unique_countries)
        sheet.cell(row=i + 3, column=39).value = unique_num_numbers
        sheet.cell(row=i + 3, column=40).value = unique_num_ids
        sheet.cell(row=i + 3, column=41).value = ', '.join(topic.most_frequent((1,1),filename='41 '+fname))
        freq_50 = topic.most_frequent(COEFFICIENT_1_FOR_NEWS, filename='42 '+fname)
        sheet.cell(row=i + 3, column=42).value = ', '.join(freq_50)

        sheet.cell(row=i + 3, column=43).value = ', '.join(topic.objects)
        sheet.cell(row=i + 3, column=44).value = ', '.join(topic.obj)
        sheet.cell(row=i + 3, column=45).value = ', '.join(topic.news[0].description.intersection(topic.news[1].description))
        sheet.cell(row=i + 3, column=46).value = ', '.join(topic.news[0].named_entities['content'].intersection(topic.news[1].named_entities['content']))

        sheet.cell(row=i + 3, column=47).value = ''
        sheet.cell(row=i + 3, column=48).value = ''
        sheet.cell(row=i + 3, column=49).value = name

        if topic.coefficient_sums:
            sheet.cell(row=i + 3, column=50).value = topic.coefficient_sums["fio_coef0"]
            sheet.cell(row=i + 3, column=51).value = topic.coefficient_sums["big_coef0"]
            sheet.cell(row=i + 3, column=52).value = topic.coefficient_sums["small_coef0"]
            sheet.cell(row=i + 3, column=53).value = topic.coefficient_sums["countries_coef0"]
            sheet.cell(row=i + 3, column=54).value = topic.coefficient_sums["fio_coefY"]
            sheet.cell(row=i + 3, column=55).value = topic.coefficient_sums["big_coefY"]
            sheet.cell(row=i + 3, column=56).value = topic.coefficient_sums["small_coefY"]
            sheet.cell(row=i + 3, column=57).value = topic.coefficient_sums["countries_coefY"]
            sheet.cell(row=i + 3, column=58).value = topic.coefficient_sums["fio_coef2"]
            sheet.cell(row=i + 3, column=59).value = topic.coefficient_sums["big_coef2"]
            sheet.cell(row=i + 3, column=60).value = topic.coefficient_sums["small_coef2"]
            sheet.cell(row=i + 3, column=61).value = topic.coefficient_sums["countries_coef2"]
            sheet.cell(row=i + 3, column=62).value = topic.coefficient_sums["ids_coef"]
            sheet.cell(row=i + 3, column=63).value = topic.coefficient_sums["summ_1"]
            sheet.cell(row=i + 3, column=64).value = topic.coefficient_sums["summ_2"]
            sheet.cell(row=i + 3, column=65).value = topic.coefficient_sums["final_result"]

        ids = {str(n.id) for n in topic.news}
        news_ids = str(len(topic.news))+' | '+', '.join(ids)

        sheet.cell(row=i + 3, column=66).value = news_ids

        col = 67

        for j in range(2, len(topic.news)):

            doc = topic.news[j]

            sheet.cell(row=i + 3, column=col).value = f"|{doc.id}| {doc.country} | {doc.url} | {doc.translated['title']} | " \
                                                      f"{doc.translated['lead']} | " \
                                                      f"{doc.translated['content']} | Из краткого: {doc.description} | Из текста: {doc.named_entities['content']}"
            try:
                sheet.cell(row=i + 3, column=col + 1).value = "| ".join(topic.methods_for_news[doc.id])
            except KeyError:
                sheet.cell(row=i + 3, column=col + 1).value = ""
            col += 2

    wb.save(fname)


def write_topics_with_subtopics(fname, topics, fio_in_freq=False, fio_in_freq_new=False):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.cell(row=1, column=1).value = "Method"
    sheet.cell(row=1, column=2).value = "Old name"  # should be empty
    sheet.cell(row=1, column=3).value = "Name"  # should be empty
    sheet.cell(row=1, column=4).value = "Topic"
    sheet.cell(row=1, column=5).value = "All unique words"

    sheet.cell(row=1, column=6).value = "News1"
    sheet.cell(row=1, column=7).value = "News2"

    sheet.cell(row=1, column=8).value = "# All"
    sheet.cell(row=1, column=9).value = "# of unique"

    sheet.cell(row=1, column=10).value = "# of All-countries"
    sheet.cell(row=1, column=11).value = "# of All-countries-small"
    sheet.cell(row=1, column=12).value = "# of FIO"
    sheet.cell(row=1, column=13).value = "# of Big letter"
    sheet.cell(row=1, column=14).value = "# of Small letter"
    sheet.cell(row=1, column=15).value = "# of Countries"
    sheet.cell(row=1, column=16).value = "# of Numbers"
    sheet.cell(row=1, column=17).value = "# of IDs"
    sheet.cell(row=1, column=18).value = "Frequent Unique"
    sheet.cell(row=1, column=19).value = "u# of All-countries (unique)"
    sheet.cell(row=1, column=20).value = "u# of unique"
    sheet.cell(row=1, column=21).value = "u# of All-countries-small (unique)"
    sheet.cell(row=1, column=22).value = "u# of FIO (unique)"
    sheet.cell(row=1, column=23).value = "u# of Big letter (unique)"
    sheet.cell(row=1, column=24).value = "u# of Small letter (unique)"
    sheet.cell(row=1, column=25).value = "u# of Countries (unique)"
    sheet.cell(row=1, column=26).value = "uNumbers (unique)"
    sheet.cell(row=1, column=27).value = "uIDs (unique)"
    sheet.cell(row=1, column=28).value = "Small letter"
    sheet.cell(row=1, column=29).value = "Countries"
    sheet.cell(row=1, column=30).value = "Numbers"
    sheet.cell(row=1, column=31).value = "IDs"
    sheet.cell(row=1, column=32).value = "uAll unique words"
    sheet.cell(row=1, column=33).value = "uAll-countries (unique)"
    sheet.cell(row=1, column=34).value = "uAll-countries-small (unique)"
    sheet.cell(row=1, column=35).value = "uFIO (unique)"
    sheet.cell(row=1, column=36).value = "uBig letter (unique)"
    sheet.cell(row=1, column=37).value = "uSmall letter (unique)"
    sheet.cell(row=1, column=38).value = "uCountries (unique)"
    sheet.cell(row=1, column=39).value = "u# of Numbers (unique)"
    sheet.cell(row=1, column=40).value = "u# of IDs (unique)"
    sheet.cell(row=1, column=41).value = "Most frequent"
    sheet.cell(row=1, column=42).value = "50% frequent"
    sheet.cell(row=1, column=43).value = "FREQUENT-NEW"
    sheet.cell(row=1, column=44).value = "What2"
    sheet.cell(row=1, column=45).value = "Common in descriptions"
    sheet.cell(row=1, column=46).value = "Common in text"
    sheet.cell(row=1, column=47).value = "Comment"  # should be empty
    sheet.cell(row=1, column=48).value = "Name"
    sheet.cell(row=1, column=49).value = "Topic"

    sheet.cell(row=1, column=50).value = "# of News, their ids"

    sheet.cell(row=1, column=51).value = "News"
    sheet.cell(row=1, column=52).value = "Method"

    # counts = [0]*7

    row_num = 3

    for i, topic in enumerate(topics):

        all_words, _ = topic.all_words(topic.name)
        all_wo_countries, num_all_wo_countries = topic.all_wo_countries(all_words)
        all_wo_countries_small, num_all_wo_countries_small = topic.all_wo_countries_and_small(all_words)
        fio, num_fio = topic.fio(all_words)
        big, num_big = topic.big(all_words)
        small, num_small = topic.small(all_words)
        countries, num_countries = topic.countries(all_words)
        numbers, num_numbers = topic.numbers(all_words)
        ids, num_ids = topic.ids(all_words)

        unique_words, num_unique_words = topic.all_words(topic.new_name)
        unique_wo_countries, num_unique_wo_countries = topic.all_wo_countries(unique_words)
        unique_wo_countries_small, num_unique_wo_countries_small = topic.all_wo_countries_and_small(unique_words)
        unique_fio, unique_num_fio = topic.fio(unique_words)
        unique_big, unique_num_big = topic.big(unique_words)
        unique_small, unique_num_small = topic.small(unique_words)
        unique_countries, unique_num_countries = topic.countries(unique_words)
        unique_numbers, unique_num_numbers = topic.numbers(unique_words)
        unique_ids, unique_num_ids = topic.ids(unique_words)
        frequent_unique = unique_words.intersection(topic.most_frequent(filename='18 '+fname))

        sheet.cell(row=row_num, column=1).value = ""
        sheet.cell(row=row_num, column=2).value = ", ".join(topic.old_name)
        sheet.cell(row=row_num, column=3).value = ""

        sheet.cell(row=row_num, column=4).value = "MAIN TOPIC: "+", ".join(topic.name)
        sheet.cell(row=row_num, column=5).value = ', '.join(unique_words)
        doc = topic.news[0]
        sheet.cell(row=row_num, column=6).value = f"|{doc.id}| {doc.country} | {doc.url} | {doc.translated['title']} | " \
                                                  f"{doc.translated['lead']} | " \
                                                  f"{doc.translated['content']} | Из краткого: {doc.description} | Из текста: {doc.named_entities['content']}"
        try:
            doc = topic.news[1]
            sheet.cell(row=row_num, column=7).value = f"|{doc.id}| {doc.country} | {doc.url} | {doc.translated['title']} | " \
                                                      f"{doc.translated['lead']} | " \
                                                      f"{doc.translated['content']} | Из краткого: {doc.description} | Из текста: {doc.named_entities['content']}"
        except IndexError:
            pass

        sheet.cell(row=row_num, column=8).value = len(all_words)
        sheet.cell(row=row_num, column=9).value = len(unique_words)

        sheet.cell(row=row_num, column=10).value = num_all_wo_countries
        sheet.cell(row=row_num, column=11).value = num_all_wo_countries_small
        sheet.cell(row=row_num, column=12).value = num_fio
        sheet.cell(row=row_num, column=13).value = num_big
        sheet.cell(row=row_num, column=14).value = num_small
        sheet.cell(row=row_num, column=15).value = num_countries
        sheet.cell(row=row_num, column=16).value = num_numbers
        sheet.cell(row=row_num, column=17).value = num_ids

        sheet.cell(row=row_num, column=18).value = ", ".join(frequent_unique)
        sheet.cell(row=row_num, column=19).value = num_unique_wo_countries
        sheet.cell(row=row_num, column=20).value = num_unique_words
        sheet.cell(row=row_num, column=21).value = num_unique_wo_countries_small
        sheet.cell(row=row_num, column=22).value = unique_num_fio
        sheet.cell(row=row_num, column=23).value = unique_num_big
        sheet.cell(row=row_num, column=24).value = unique_num_small
        sheet.cell(row=row_num, column=25).value = unique_num_countries
        sheet.cell(row=row_num, column=26).value = ", ".join(unique_numbers)
        sheet.cell(row=row_num, column=27).value = ", ".join(unique_ids)
        sheet.cell(row=row_num, column=28).value = ", ".join(small)
        sheet.cell(row=row_num, column=29).value = ", ".join(countries)
        sheet.cell(row=row_num, column=30).value = ", ".join(numbers)
        sheet.cell(row=row_num, column=31).value = ", ".join(ids)
        sheet.cell(row=row_num, column=32).value = ", ".join(unique_words)
        sheet.cell(row=row_num, column=33).value = ", ".join(unique_wo_countries)
        sheet.cell(row=row_num, column=34).value = ", ".join(unique_wo_countries_small)
        sheet.cell(row=row_num, column=35).value = ", ".join(unique_fio)
        sheet.cell(row=row_num, column=36).value = ", ".join(unique_big)
        sheet.cell(row=row_num, column=37).value = ", ".join(unique_small)
        sheet.cell(row=row_num, column=38).value = ", ".join(unique_countries)
        sheet.cell(row=row_num, column=39).value = unique_num_numbers
        sheet.cell(row=row_num, column=40).value = unique_num_ids

        sheet.cell(row=row_num, column=41).value = ', '.join(topic.most_frequent((1, 1), fio_in_freq, filename='41 '+fname))
        freq_50 = topic.most_frequent(COEFFICIENT_1_FOR_NEWS, fio_in_freq, filename='42 '+fname)

        sheet.cell(row=row_num, column=42).value = ', '.join(freq_50)

        sheet.cell(row=row_num, column=43).value = ', '.join(topic.most_frequent(COEFFICIENT_2_FOR_NEWS, fio_in_freq_new, filename='43 '+fname))
        sheet.cell(row=row_num, column=44).value = ', '.join(topic.obj)
        try:
            sheet.cell(row=row_num, column=45).value = ', '.join(
                topic.news[0].description.intersection(topic.news[1].description))
            sheet.cell(row=row_num, column=46).value = ', '.join(
                topic.news[0].named_entities['content'].intersection(topic.news[1].named_entities['content']))
        except IndexError:
            pass

        sheet.cell(row=row_num, column=47).value = ''
        sheet.cell(row=row_num, column=48).value = ''
        sheet.cell(row=row_num, column=49).value = ', '.join(topic.name)

        ids = {str(n.id) for n in topic.news}
        news_ids = str(len(topic.news)) + ' | ' + ', '.join(ids)

        sheet.cell(row=row_num, column=50).value = news_ids

        # sheet.cell(row=i + 3, column=51).value = ', '.join(topic.method)
        # sheet.cell(row=i + 3, column=50).value = ', '.join(topic.all_numbers)

        col = 51

        for j in range(2, len(topic.news)):

            doc = topic.news[j]

            sheet.cell(row=row_num,
                       column=col).value = f"|{doc.id}| {doc.country} | {doc.url} | {doc.translated['title']} | " \
                                           f"{doc.translated['lead']} | " \
                                           f"{doc.translated['content']} | Из краткого: {doc.description} | Из текста: {doc.named_entities['content']}"
            try:
                sheet.cell(row=row_num, column=col + 1).value = "| ".join(topic.methods_for_news[doc.id])
            except KeyError:
                sheet.cell(row=row_num, column=col + 1).value = ""
            col += 2

        if topic.subtopics:
            for k, s in enumerate(topic.subtopics):
                all_words, _ = s.all_words(s.name)
                all_wo_countries, num_all_wo_countries = s.all_wo_countries(all_words)
                all_wo_countries_small, num_all_wo_countries_small = s.all_wo_countries_and_small(all_words)
                fio, num_fio = s.fio(all_words)
                big, num_big = s.big(all_words)
                small, num_small = s.small(all_words)
                countries, num_countries = s.countries(all_words)
                numbers, num_numbers = s.numbers(all_words)
                ids, num_ids = s.ids(all_words)

                unique_words, num_unique_words = s.all_words(s.new_name)
                unique_wo_countries, num_unique_wo_countries = s.all_wo_countries(unique_words)
                unique_wo_countries_small, num_unique_wo_countries_small = s.all_wo_countries_and_small(unique_words)
                unique_fio, unique_num_fio = s.fio(unique_words)
                unique_big, unique_num_big = s.big(unique_words)
                unique_small, unique_num_small = s.small(unique_words)
                unique_countries, unique_num_countries = s.countries(unique_words)
                unique_numbers, unique_num_numbers = s.numbers(unique_words)
                unique_ids, unique_num_ids = s.ids(unique_words)
                frequent_unique = unique_words.intersection(s.most_frequent())
                
                row_num += 1

                sheet.cell(row=row_num, column=1).value = " ".join(s.method)
                sheet.cell(row=row_num, column=2).value = ", ".join(s.old_name)
                sheet.cell(row=row_num, column=3).value = ""

                sheet.cell(row=row_num, column=4).value = ", ".join(s.name)
                sheet.cell(row=row_num, column=5).value = ', '.join(unique_words)
                doc = s.news[0]
                sheet.cell(row=row_num,
                           column=6).value = f"|{doc.id}| {doc.country} | {doc.url} | {doc.translated['title']} | " \
                                             f"{doc.translated['lead']} | " \
                                             f"{doc.translated['content']} | Из краткого: {doc.description} | Из текста: {doc.named_entities['content']}"
                try:
                    doc = s.news[1]
                    sheet.cell(row=row_num,
                               column=7).value = f"|{doc.id}| {doc.country} | {doc.url} | {doc.translated['title']} | " \
                                                 f"{doc.translated['lead']} | " \
                                                 f"{doc.translated['content']} | Из краткого: {doc.description} | Из текста: {doc.named_entities['content']}"
                except IndexError:
                    pass

                sheet.cell(row=row_num, column=8).value = len(all_words)
                sheet.cell(row=row_num, column=9).value = len(unique_words)

                sheet.cell(row=row_num, column=10).value = num_all_wo_countries
                sheet.cell(row=row_num, column=11).value = num_all_wo_countries_small
                sheet.cell(row=row_num, column=12).value = num_fio
                sheet.cell(row=row_num, column=13).value = num_big
                sheet.cell(row=row_num, column=14).value = num_small
                sheet.cell(row=row_num, column=15).value = num_countries
                sheet.cell(row=row_num, column=16).value = num_numbers
                sheet.cell(row=row_num, column=17).value = num_ids

                sheet.cell(row=row_num, column=18).value = ", ".join(frequent_unique)
                sheet.cell(row=row_num, column=19).value = num_unique_wo_countries
                sheet.cell(row=row_num, column=20).value = num_unique_words
                sheet.cell(row=row_num, column=21).value = num_unique_wo_countries_small
                sheet.cell(row=row_num, column=22).value = unique_num_fio
                sheet.cell(row=row_num, column=23).value = unique_num_big
                sheet.cell(row=row_num, column=24).value = unique_num_small
                sheet.cell(row=row_num, column=25).value = unique_num_countries
                sheet.cell(row=row_num, column=26).value = ", ".join(unique_numbers)
                sheet.cell(row=row_num, column=27).value = ", ".join(unique_ids)
                sheet.cell(row=row_num, column=28).value = ", ".join(small)
                sheet.cell(row=row_num, column=29).value = ", ".join(countries)
                sheet.cell(row=row_num, column=30).value = ", ".join(numbers)
                sheet.cell(row=row_num, column=31).value = ", ".join(ids)
                sheet.cell(row=row_num, column=32).value = ", ".join(unique_words)
                sheet.cell(row=row_num, column=33).value = ", ".join(unique_wo_countries)
                sheet.cell(row=row_num, column=34).value = ", ".join(unique_wo_countries_small)
                sheet.cell(row=row_num, column=35).value = ", ".join(unique_fio)
                sheet.cell(row=row_num, column=36).value = ", ".join(unique_big)
                sheet.cell(row=row_num, column=37).value = ", ".join(unique_small)
                sheet.cell(row=row_num, column=38).value = ", ".join(unique_countries)
                sheet.cell(row=row_num, column=39).value = unique_num_numbers
                sheet.cell(row=row_num, column=40).value = unique_num_ids

                sheet.cell(row=row_num, column=41).value = ', '.join(s.most_frequent((1, 1), fio_in_freq))
                freq_50 = s.most_frequent(COEFFICIENT_1_FOR_NEWS, fio_in_freq)

                sheet.cell(row=row_num, column=42).value = ', '.join(freq_50)

                sheet.cell(row=row_num, column=43).value = ', '.join(s.most_frequent(COEFFICIENT_2_FOR_NEWS, fio_in_freq_new))
                sheet.cell(row=row_num, column=44).value = ', '.join(s.obj)
                try:
                    sheet.cell(row=row_num, column=45).value = ', '.join(
                        s.news[0].description.intersection(s.news[1].description))
                    sheet.cell(row=row_num, column=46).value = ', '.join(
                        s.news[0].named_entities['content'].intersection(s.news[1].named_entities['content']))
                except IndexError:
                    pass

                sheet.cell(row=row_num, column=47).value = ''
                sheet.cell(row=row_num, column=48).value = ''
                sheet.cell(row=row_num, column=49).value = ', '.join(s.name)
                sheet.cell(row=row_num, column=50).value = len(s.news)

                # sheet.cell(row=i + 3, column=51).value = ', '.join(s.method)
                # sheet.cell(row=i + 3, column=50).value = ', '.join(s.all_numbers)

                col = 51

                for j in range(2, len(s.news)):

                    doc = s.news[j]

                    sheet.cell(row=row_num,
                               column=col).value = f"|{doc.id}| {doc.country} | {doc.url} | {doc.translated['title']} | " \
                                                   f"{doc.translated['lead']} | " \
                                                   f"{doc.translated['content']} | Из краткого: {doc.description} | Из текста: {doc.named_entities['content']}"
                    try:
                        sheet.cell(row=row_num, column=col + 1).value = "| ".join(s.methods_for_news[doc.id])
                    except KeyError:
                        sheet.cell(row=row_num, column=col + 1).value = ""
                    col += 2

        row_num += 2

    wb.save(fname)


def write_news(fname, news):
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i, doc in enumerate(news):
        sheet.cell(row=i + 1, column=1).value = f"{doc.id}"
        sheet.cell(row=i + 1, column=2).value = f"{doc.country}"
        sheet.cell(row=i + 1, column=3).value = f"{doc.url}"

        sheet.cell(row=i + 1, column=4).value = f"{doc.translated['title']} | {doc.named_entities['title']}"
        sheet.cell(row=i + 1, column=5).value = f"{doc.translated['lead']} | {doc.named_entities['lead']}"
        sheet.cell(row=i + 1, column=6).value = f"{doc.translated['content']} | {doc.named_entities['content']}"

    wb.save(fname)