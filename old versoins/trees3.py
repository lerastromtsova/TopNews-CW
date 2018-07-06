import re
import time

import nltk
import sqlite3
import csv
import openpyxl

from text_processing.preprocess import preprocess
from text_processing.translate import translate
from text_processing.dates import process_dates


class Corpus:

    def __init__(self, db, table):

        self.db = db
        self.table = table
        self.conn = sqlite3.connect(f"db/{db}.db")
        self.conn.row_factory = sqlite3.Row
        self.c = self.conn.cursor()
        self.c.execute("SELECT * FROM " + table)
        self.topics = []
        self.data = []

        raw_data = self.c.fetchall()

        for row in raw_data:
            doc = Document(row, self.conn, table)
            doc.process(['title','lead','content'])
            self.data.append(doc)

    def find_topics(self):

        for row1 in self.data:
            rows_except_this = [r for r in self.data if r.url != row1.url and r.country != row1.country]
            for row2 in rows_except_this:
                com_words = row1.named_entities['content'].intersection(row2.named_entities['content'])
                if len(com_words) and com_words not in self.topics:
                    self.topics.append(com_words)

        self.topics.sort(key=lambda s: -len(s))


class Document:

    def __init__(self, row, conn, table):

        types = ('title', 'lead', 'content')
        self.raw = row

        self.country = row['country']
        self.date = row['date']
        self.orig_data = dict.fromkeys(types)
        self.orig_data['title'] = row['title']

        try:
            self.orig_data['lead'] = row['lead']
        except IndexError:
            self.orig_data['lead'] = row['description']

        self.orig_data['content'] = row['content']
        self.url = row['reference']

        self.translated = dict.fromkeys(types)
        self.double_translated = dict.fromkeys(types)
        self.tokens = dict.fromkeys(types)
        self.named_entities = dict.fromkeys(types)

        self.conn = conn
        self.table = table

        self.dates = process_dates(list(self.tokens)).append(self.date)

    def process(self, arr_of_types):

        for typ in arr_of_types:

            if typ == 'content':
                col = 'translated'
                col1 = 'translated1'

            else:
                col = f'translated_{typ}'
                col1 = f'translated1_{typ}'

            if self.raw[col]:
                self.translated[typ] = self.raw[col]
            else:
                self.double_translate(typ)

            if self.raw[col1]:
                self.double_translated[typ] = self.raw[col1]
            else:
                self.double_translate(typ)

            self.tokens[typ] = [word for word in preprocess(self.translated[typ]) if
                                word in preprocess(self.double_translated[typ])]

            parse_tree = nltk.ne_chunk(nltk.tag.pos_tag(self.tokens[typ]),
                                       binary=True)  # POS tagging before chunking!

            self.named_entities[typ] = {k[0] for branch in parse_tree.subtrees() for k in list(branch) if branch.label() == 'NE'}
            self.unite_countries_in(typ,'nes')
            self.find_entities(typ)
            self.unite_countries_in(typ,'nes')

    def find_entities(self, ty):

        text = re.findall(r"[\w]+|[^\s\w]", self.translated[ty])

        to_remove = set()
        to_add = set()

        nes = self.named_entities[ty]

        for ent1 in nes:
            if ent1 in text:
                idx1 = text.index(ent1)
                entities_except_this = nes - set(ent1)

                for ent2 in entities_except_this:
                    if ent2 in text:
                        idx2 = text.index(ent2)
                        if ((idx2 - idx1 == 2) and (text[idx1+1] == ' ' or text[idx1+1] == '-'
                                                    or text[idx1+1] == "'" or text[idx1+1] == 'of')) or idx2 - idx1 == 1:

                            united_entity = ' '.join([ent1, ent2])
                            to_add.add(united_entity)
                            to_remove.add(ent1)
                            to_remove.add(ent2)

        self.named_entities[ty] = (self.named_entities[ty] - to_remove) | to_add

    def unite_countries_in(self, ty, type_of_data):
        conn = sqlite3.connect("db/countries.db")
        c = conn.cursor()
        c.execute("SELECT * FROM countries")
        all_rows = c.fetchall()
        to_remove = set()
        to_add = set()
        if type_of_data == "nes":
            data = self.named_entities[ty]
        elif type_of_data=='tokens':
            data = self.tokens[ty]
        for ent in data:
                for row in all_rows:
                    low = [w.lower() for w in row if w is not None]
                    if ent.lower() in low:
                        to_remove.add(ent)
                        to_add.add(row[0])

        if type_of_data == 'nes':
            self.named_entities[ty] = (self.named_entities[ty] - to_remove) | to_add
        elif type_of_data == 'tokens':
            self.tokens[ty].extend(list(to_add))
            self.tokens[ty] = [t for t in self.tokens[ty] if t not in to_remove]

    def double_translate(self, ty):

        n = 1500  # length limit
        text = self.orig_data[ty]
        self.translated[ty] = ''
        self.double_translated[ty] = ''

        if "Краткое описание: " in text:
            text = text.split("Краткое описание: ")[1]

        # Split into parts of 1500 before translating
        text = [text[i:i + n] for i in range(0, len(text), n)]

        for part in text:

            eng_text = translate(part)
            orig_text = translate(eng_text, self.country)
            eng1_text = translate(orig_text)

            self.translated[ty] += ' '
            self.translated[ty] += eng_text
            self.double_translated[ty] += ' '
            self.double_translated[ty] += eng1_text

        c = self.conn.cursor()
        if ty == 'content':
            col = 'translated'
            col1 = 'translated1'

        else:
            col = f'translated_{ty}'
            col1 = f'translated1_{ty}'

        c.execute(f"UPDATE {self.table} SET {col}=(?), {col1}=(?) WHERE reference=(?)",
                  (self.translated[ty], self.double_translated[ty], self.url))
        self.conn.commit()


class Node:

    def __init__(self, item):

        self.name = item
        self.level = len(self.name)

        self.parents = []
        self.children = []
        self.documents = []
        self.percents = []
        self.subtopics = []

    def has_free_links(self, other_node):

        topic = self.name
        my_free_words = [d.named_entities['content'] - topic for d in self.documents]
        suit_documents = [sd for sd in self.documents if sd not in other_node.documents]
        par_free_words = [pd.named_entities['content'] - topic for pd in suit_documents]
        connection = 0

        for doc in my_free_words:
            for d in par_free_words:
                if doc.intersection(d):
                    connection += 1

        if my_free_words and par_free_words:
            percent_of_connected = connection/(len(my_free_words)*len(par_free_words))
            return percent_of_connected
        return 0

    def find_most_frequent(self, type):
        freq_words = {}
        for document in self.documents:

            # descr = []
            # descr.extend(document.tokens['title'])
            # descr.extend(document.tokens['lead'])lead
            descr = set(document.tokens['title']) | set(document.tokens['lead'])

            if type == 'lowercase':
                parse_tree = nltk.ne_chunk(nltk.tag.pos_tag(descr),
                                           binary=True)  # POS tagging before chunking!

                nes = {k[0] for branch in parse_tree.subtrees() for k in list(branch) if
                                            branch.label() == 'NE'}
                descr = [d for d in descr if d not in nes]

            for word in descr:

                countries = {d.country for d in self.documents if word in d.tokens['lead'] or word in d.tokens['title']}
                if len(countries) >= 2:

                    if word in freq_words.keys():
                        freq_words[word] += 1
                    else:
                        freq_words[word] = 1

        if freq_words:
            maxval = max(freq_words.values())
            result = {key: value for key, value in freq_words.items() if value == maxval}
            return result
        else:
            return freq_words

    def add_document(self, doc):
        self.documents.append(doc)

    def assign_parent(self, parent):
        if parent not in self.parents:
            self.parents.append(parent)
            self.percents.append(self.has_free_links(parent))

    def assign_child(self, child):
        if child not in self.children:
            self.children.append(child)

    def isparent(self, other_node):
        if other_node in self.parents:
            return True
        return False

    def ischild(self, other_node):
        if other_node in self.children:
            return True
        return False

    def isroot(self):
        if self.parents:
            return False
        return True

    def isleaf(self):
        if self.children:
            return False
        return True

    def contains(self, children):
        for ch in children:
            if self.name.issubset(ch.name):
                return True
        return False


class Tree:

    def __init__(self, corpus):

        self.nodes = [Node(item) for item in corpus.topics]
        self.assign_documents(corpus)
        self.db = corpus.db

        for node in self.nodes:

            for other_node in [n for n in self.nodes if n != node]:

                if node.name.issubset(other_node.name) and node not in other_node.children and not node.contains(other_node.children):
                    node.assign_parent(other_node)
                    other_node.assign_child(node)

                if other_node.name.issubset(node.name) and node not in other_node.parents and not other_node.contains(node.children):
                    other_node.assign_parent(node)
                    node.assign_child(other_node)

        self.roots = [n for n in self.nodes if n.isroot()]
        self.leaves = [n for n in self.nodes if n.isleaf()]
        write_topics_to_xl("0.xlsx", self.roots)
        self.last_nodes = set()
        # self.last_nodes = set(self.roots)

    def assign_documents(self, corpus):

        for node in self.nodes:
            for doc in corpus.data:
                if node.name.issubset(doc.named_entities['content']):
                    node.add_document(doc)

    def find_last_topics(self):

        for node in self.roots:
            name, docs = get_all_in_strings(node, [], set())
            new_node = Node(name)
            new_node.documents = docs
            self.last_nodes.add(new_node)
        write_topics_to_xl("3-2.xlsx", self.last_nodes)
        write_tz(self)


    def unite_similar_topics(self):

        to_remove = set()

        for n in self.roots:
            frequent_words_dict = n.find_most_frequent(type='uppercase')

            if any(k < 0.5 * len(n.documents) for k in frequent_words_dict.values()):
                print("Removed: ", n.name)
                to_remove.add(n)
            elif not frequent_words_dict.values():
                print("Removed: ", n.name)
                to_remove.add(n)
            else:
                print("Not removed: ",n.name)
                has_these_words = []
                print(frequent_words_dict)

                for d in n.documents:
                    for w in frequent_words_dict.keys():
                        if w in d.tokens['title'] or w in d.tokens['lead']:
                            has_these_words.append(d)
                            continue

                rest = [d for d in n.documents if d not in has_these_words]
                print(len(has_these_words))
                print(len(rest))

                for doc in rest:
                    perc = 0
                    for doc1 in has_these_words:
                        common_words = set(doc.tokens['title']).intersection(set(doc1.tokens['title']))
                        potential_words = common_words - set(frequent_words_dict.keys())
                        if not potential_words:
                            if doc in n.documents:
                                n.documents.remove(doc)

            countries = {d.country for d in n.documents}
            if len(countries) == 1:
                to_remove.add(n)

        self.roots = [r for r in self.roots if r not in to_remove]
        write_topics_with_freq("3-1.xlsx", self.roots, with_children=False)

    def unite_by_countries(self):
        most_similar = []

        while self.last_nodes:
            node = self.last_nodes.pop()
            others = self.last_nodes - {node}
            similar = [node]
            while others:
                other = others.pop()

                common_docs = [nd for nd in node.documents for od in other.documents if
                               nd.translated['title'] == od.translated['title']]
                common_countries = {cd.country for cd in common_docs}
                if len(common_countries) >= 2:
                    self.last_nodes.remove(other)
                    similar.append(other)
                    print(node.name, other.name)
            most_similar.append(similar)

        for ms in most_similar:
            name = set()
            docs = set()
            subtopics = []
            for s in ms:
                name |= set(s.name)
                docs |= set(s.documents)
                subtopics.append(s)

            node = Node(name)
            node.documents = docs
            node.subtopics = subtopics
            self.last_nodes.add(node)

        write_topics_to_xl("3-3.xlsx", self.last_nodes, with_children=True)

    def delete_subsets(self):
        to_remove = set()

        for n in self.last_nodes:
            other_nodes = self.last_nodes - {n}
            for on in other_nodes:
                if n.name.issubset(on.name) and on not in to_remove:
                    to_remove.add(n)

        return to_remove

    def add_and_remove(self, principle):
        all_links = []
        to_remove = set()
        to_add = set()

        for node in self.last_nodes:
            countries = {d.country for d in node.documents}
            print(node.name)
            print(countries)
            nodes_except_this = self.last_nodes - {node}

            for other_node in nodes_except_this:

                other_countries = {d.country for d in other_node.documents}

                common_documents = set(node.documents).intersection(set(other_node.documents))
                common_countries = {d.country for d in common_documents}
                percent_1 = len(common_countries) / len(countries)
                percent_2 = len(common_countries) / len(other_countries)

                if percent_1 > 0.5 and percent_2 > 0.5:
                        all_links.append((node, other_node, percent_1*percent_2))

        for node in self.last_nodes:
            try:
                l = []
                for a in all_links:
                    if a[0] == node:
                        l.append((a[1], a[2]))

                max_value = max(l, key=lambda x: x[1])[1]
                max_sim_nodes = [s[0] for s in l if s[1] == max_value]
                name = node.name
                subtopics = []
                documents = node.documents
                to_remove.add(node)
                print(f"Node: {node.name}")

                for msn in max_sim_nodes:
                    print(f"Max_sim_node: {msn.name}")
                    name |= msn.name
                    subtopics.append(msn)

                    documents.extend(msn.documents)

                    to_remove.add(msn)

                new_node = Node(name)
                new_node.subtopics = subtopics
                new_node.documents = list(set(documents))
                to_add.add(new_node)

                # new_node.children = [node, maximum_similarity_node]
            except ValueError:
                continue

        return to_remove, to_add


def check_free(nodes):
    for node in nodes:
        for ch in node.children:
            ch_percent = ch.percents[ch.parents.index(node)]
            for gch in ch.children:
                gchild_percent = gch.percents[gch.parents.index(ch)]
                if gchild_percent < ch_percent and gchild_percent < 0.5:
                    return ch
                elif gch.children:
                    return check_free(ch.children)
                else:
                    return gch


def check_xl(nodes, all_links):
    file = csv.writer(open("t.csv", "w"), delimiter=',')

    for n in nodes:
        countries = {d.country for d in n.documents}
        row = [f"{n.name} | {countries}"]
        for k, m in all_links[n].items():
            row.append(f"{k.name} | {m} | {' '.join({d.country for d in k.documents})}")
            cd = set(n.documents).intersection(set(k.documents))
            cc = {d.country for d in cd}
            perc = len(cc)/len(countries)
            row.append(perc)
            row.append(cc)
            tex = ' | '.join([d.title for d in cd])
            row.append(tex)
        file.writerow(row)


def write_topics_with_freq(fname, nodes, with_children=False):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = 'Topic'
    sheet['B1'] = 'News'
    sheet['C1'] = 'Keywords'
    for i, n in enumerate(nodes):
        if with_children:
            text = ''
            if n.subtopics:
                for c in n.subtopics:
                    text += ' {'.join(c.name)
                    text += ' | '
            else:
                text = ' '.join(n.name)
        else:
            text = ' '.join(n.name)
        freq_words = n.find_most_frequent(type="uppercase")
        for word, val in freq_words.items():
            t = f"|{word}:{val}|"
            text += t
        sheet.cell(row=i+2, column=1).value = text
        docs = n.documents
        for j, doc in enumerate(docs):
            sheet.cell(row=i+2, column=j+2).value = f"{doc.country} | {doc.translated['title']} | " \
                                                    f"{doc.translated['lead']} | {doc.url} | {doc.translated['content']} " \
                                                    f"| {doc.named_entities['title']} | {doc.named_entities['content']}"
    wb.save(fname)


def write_topics_to_xl(fname, nodes, with_children=False):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = 'Topic'
    sheet['B1'] = 'News'
    sheet['C1'] = 'Keywords'
    for i, n in enumerate(nodes):
        if with_children:
            text = ''
            if n.subtopics:
                for c in n.subtopics:
                    text += ' {'.join(c.name)
                    text += ' | '
            else:
                text = ' '.join(n.name)
        else:
            text = ' '.join(n.name)
        sheet.cell(row=i+2, column=1).value = text
        docs = n.documents
        for j, doc in enumerate(docs):
            sheet.cell(row=i+2, column=j+2).value = f"{doc.country} | {doc.translated['title']} | {doc.translated['lead']} | {doc.url} | {doc.translated['content']} | {doc.named_entities['title']} | {doc.named_entities['content']}"
    wb.save(fname)


def write_tz(tree):
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i, node in enumerate(tree.roots):
        texts, _ = get_all_in_strings(node, [], set())
        for j, text in enumerate(texts):
            sheet.cell(row=i+1, column=j+1).value = text
    wb.save("tz-today.xlsx")


def get_all_in_strings(node, prev, docs):
    docs.update(node.documents)
    prev.append(f"{node.name} | {len(node.documents)}")
    for ch in node.children:
        if ch.children:
            get_all_in_strings(ch, prev, docs)
    return prev, docs


def write_words_to_xl(fname, data):
    file = csv.writer(open(fname, "w"), delimiter=',')
    headers = ['Document', 'Deleted words']
    file.writerow(headers)

    for r in data:
        row = [r.translated, r.removed_words]
        file.writerow(row)


def write_start_words_to_xl(fname, data):
    file = csv.writer(open(fname, "w"))
    headers = ['Document', 'Deleted words']
    file.writerow(headers)

    for r in data:
        for key, ind in r.start_words.items():
            row = [key, ind]
            file.writerow(row)


if __name__ == '__main__':

    now = time.time()

    db = input("DB name (default - day): ")
    table = input("Table name (default - buffer): ")

    if not db:
        db = "day"
    if not table:
        table = "buffer"

    c = Corpus(db, table)
    c.find_topics()
    # write_words_to_xl("double_translation.csv", c.data)
    # write_start_words_to_xl("start_words.csv", c.data)

    t = Tree(c)
    t.unite_similar_topics()
    t.find_last_topics()
    t.unite_by_countries()
    # write_topics_to_xl(f"{db}-topics.csv", t.last_nodes)
    f = open("time1.txt", "w")
    f.write(str(time.time()-now))